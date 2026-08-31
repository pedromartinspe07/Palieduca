import os
import re
import shutil
import uuid
import json
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form, Request, Header, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel
import httpx
from dotenv import load_dotenv
from sqlalchemy.orm import Session
from sqlalchemy import text

from database import engine, get_db
import models
import schemas
import auth

import seed
import email_service
import moderation_bot

load_dotenv()

models.Base.metadata.create_all(bind=engine)

# Migração simples para SQLite (se as colunas não existirem, adiciona individualmente)
with engine.connect() as conn:
    for migration_sql in [
        "ALTER TABLE page_content ADD COLUMN draft_content VARCHAR",
        "ALTER TABLE page_content ADD COLUMN meta_title VARCHAR",
        "ALTER TABLE page_content ADD COLUMN meta_description VARCHAR",
        "ALTER TABLE page_content ADD COLUMN slug VARCHAR",
        "ALTER TABLE users ADD COLUMN email_verified BOOLEAN DEFAULT 0",
        "ALTER TABLE users ADD COLUMN verification_code VARCHAR",
        "ALTER TABLE users ADD COLUMN reset_password_code VARCHAR",
        "ALTER TABLE users ADD COLUMN auth_provider VARCHAR DEFAULT 'local'",
        "ALTER TABLE users ADD COLUMN last_password_change VARCHAR",
        "ALTER TABLE users ADD COLUMN foto_url VARCHAR",
        "ALTER TABLE users ADD COLUMN completion_email_sent BOOLEAN DEFAULT 0",
    ]:
        try:
            conn.execute(text(migration_sql))
            conn.commit()
        except Exception:
            pass

# Auto-seed inicial para garantir dados no Render mesmo em ambientes voláteis
seed.seed_users()
seed.seed_modules()
seed.seed_pages()

app = FastAPI()

# ==========================================
# Cibersegurança: Configurações de Upload
# ==========================================
ALLOWED_MEDIA_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg", ".pdf", ".mp3", ".mp4", ".wav"}
ALLOWED_AVATAR_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg"}
MAX_MEDIA_FILE_SIZE = 25 * 1024 * 1024  # 25 MB
MAX_AVATAR_FILE_SIZE = 5 * 1024 * 1024  # 5 MB

def sanitize_filename(filename: str) -> str:
    """Remove caracteres perigosos, espaços e sequências de path traversal."""
    base_name = os.path.basename(filename)
    clean_name = re.sub(r'[^a-zA-Z0-9_.-]', '_', base_name)
    return clean_name[:80] or "arquivo"

# Configura pasta para upload de mídias
os.makedirs("static/uploads", exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Permitir requisições tanto locais quanto do seu site oficial
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "https://palieduca.com.br",
        "https://www.palieduca.com.br",
        "https://pedromartinspe07.github.io"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    payload = auth.decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token inválido ou expirado")
    
    email = payload.get("sub")
    user = db.query(models.User).filter(models.User.email == email).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Usuário não encontrado")
    return user

class ChatMessage(BaseModel):
    role: str
    content: str

class ChatRequest(BaseModel):
    messages: list[ChatMessage]

SYSTEM_PROMPT = {
    "role": "system",
    "content": "Você é um assistente virtual especialista em Enfermagem e Cuidados Paliativos. Seu tom deve ser sempre acolhedor, empático, profissional e pautado na ética de enfermagem. Suas respostas devem focar no alívio do sofrimento, controle de sintomas (dor, dispneia, etc.) e apoio psicológico ao paciente e à família. Se o usuário perguntar sobre tópicos fora da área da saúde ou cuidados paliativos, redirecione educadamente a conversa de volta para o seu propósito principal. Não prescreva medicamentos, apenas discuta manejos baseados em evidências."
}

@app.get("/api/health")
async def health_check():
    return {"status": "ok"}

@app.post("/api/auth/register")
def register_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    db_user = db.query(models.User).filter(models.User.email == user.email).first()
    if db_user:
        if not db_user.email_verified:
            # Reenvia código se ainda não confirmou
            code = email_service.generate_verification_code()
            db_user.verification_code = code
            db.commit()
            email_service.send_verification_email(db_user.email, db_user.nome, code)
            return {
                "require_verification": True,
                "email": db_user.email,
                "message": "Conta já cadastrada aguardando verificação. Novo código enviado para o seu e-mail!"
            }
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    
    hashed_password = auth.get_password_hash(user.senha)
    code = email_service.generate_verification_code()
    
    # Se for conta dona ou desenvolvedor já nasce verificada
    is_admin = user.cargo in ["dona", "desenvolvedor"]
    
    new_user = models.User(
        email=user.email,
        nome=user.nome,
        senha_hash=hashed_password,
        cargo=user.cargo or "aluno",
        email_verified=is_admin,
        verification_code=None if is_admin else code,
        auth_provider="local"
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    if not is_admin:
        email_service.send_verification_email(new_user.email, new_user.nome, code)
        return {
            "require_verification": True,
            "email": new_user.email,
            "message": f"Enviamos um código de confirmação de 6 dígitos para {new_user.email}."
        }
    
    # Se for admin, já loga direto
    access_token = auth.create_access_token(data={"sub": new_user.email, "role": new_user.cargo})
    return {
        "require_verification": False,
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": new_user.id,
            "email": new_user.email,
            "nome": new_user.nome,
            "cargo": new_user.cargo
        }
    }

@app.post("/api/auth/verify-email")
def verify_email(data: schemas.VerifyEmailRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == data.email).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")
    
    if user.email_verified:
        access_token = auth.create_access_token(data={"sub": user.email, "role": user.cargo})
        return {
            "message": "E-mail já verificado!",
            "access_token": access_token,
            "token_type": "bearer",
            "user": {
                "id": user.id,
                "email": user.email,
                "nome": user.nome,
                "cargo": user.cargo
            }
        }
    
    if not user.verification_code or user.verification_code.strip() != data.code.strip():
        raise HTTPException(status_code=400, detail="Código de verificação incorreto ou expirado.")
    
    user.email_verified = True
    user.verification_code = None
    db.commit()
    db.refresh(user)
    
    access_token = auth.create_access_token(data={"sub": user.email, "role": user.cargo})
    return {
        "message": "E-mail verificado com sucesso!",
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "email": user.email,
            "nome": user.nome,
            "cargo": user.cargo,
            "foto_url": user.foto_url
        }
    }

# In-memory Rate Limiter para Ciberseguranca (Protecao Anti-Brute Force e Anti-Spam)
RATE_LIMIT_STORE: dict[str, list[float]] = {}

def check_rate_limit(key: str, max_requests: int, window_seconds: int) -> bool:
    now = datetime.now().timestamp()
    timestamps = RATE_LIMIT_STORE.get(key, [])
    # Filtra apenas timestamps dentro da janela
    timestamps = [t for t in timestamps if now - t < window_seconds]
    if len(timestamps) >= max_requests:
        return False
    timestamps.append(now)
    RATE_LIMIT_STORE[key] = timestamps
    return True

@app.post("/api/auth/resend-code")
def resend_verification_code(data: schemas.ResendCodeRequest, request: Request, db: Session = Depends(get_db)):
    client_ip = request.client.host if request.client else "unknown"
    if not check_rate_limit(f"resend_{data.email}_{client_ip}", max_requests=3, window_seconds=300):
        raise HTTPException(
            status_code=429,
            detail="Muitas solicitações de código seguidas. Por segurança, aguarde 5 minutos antes de tentar novamente."
        )

    user = db.query(models.User).filter(models.User.email == data.email).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")
    
    if user.email_verified:
        return {"message": "Este e-mail já está verificado."}
    
    code = email_service.generate_verification_code()
    user.verification_code = code
    db.commit()
    
    email_service.send_verification_email(user.email, user.nome, code)
    return {"message": f"Novo código enviado para {user.email}!"}

@app.post("/api/auth/forgot-password")
def forgot_password(data: schemas.ForgotPasswordRequest, request: Request, db: Session = Depends(get_db)):
    client_ip = get_client_ip(request)
    if not check_rate_limit(f"forgot_{data.email}_{client_ip}", max_requests=4, window_seconds=300):
        raise HTTPException(
            status_code=429,
            detail="Muitas tentativas de recuperação de senha. Por segurança, aguarde 5 minutos antes de tentar novamente."
        )

    user = db.query(models.User).filter(models.User.email == data.email).first()
    if not user:
        # Resposta genérica para segurança contra enumeração
        return {"message": "Se o e-mail informado estiver cadastrado, enviamos um código de 6 dígitos para redefinição da sua senha."}

    code = email_service.generate_verification_code()
    user.reset_password_code = code
    db.commit()

    email_service.send_password_reset_email(user.email, user.nome, code)
    return {"message": "Se o e-mail informado estiver cadastrado, enviamos um código de 6 dígitos para redefinição da sua senha."}

@app.post("/api/auth/reset-password")
def reset_password(data: schemas.ResetPasswordRequest, request: Request, db: Session = Depends(get_db)):
    client_ip = get_client_ip(request)
    if not check_rate_limit(f"reset_{data.email}_{client_ip}", max_requests=5, window_seconds=300):
        raise HTTPException(
            status_code=429,
            detail="Muitas tentativas de redefinição de senha. Por segurança, aguarde 5 minutos."
        )

    user = db.query(models.User).filter(models.User.email == data.email).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")

    if not user.reset_password_code or user.reset_password_code.strip() != data.code.strip():
        raise HTTPException(status_code=400, detail="Código de recuperação incorreto ou expirado.")

    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="A nova senha deve ter pelo menos 6 caracteres.")

    user.senha_hash = auth.get_password_hash(data.new_password)
    user.reset_password_code = None
    user.last_password_change = datetime.now().isoformat()
    user.email_verified = True
    db.commit()
    db.refresh(user)

    access_token = auth.create_access_token(data={"sub": user.email, "role": user.cargo})
    return {
        "message": "Senha redefinida com sucesso!",
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "email": user.email,
            "nome": user.nome,
            "cargo": user.cargo,
            "foto_url": user.foto_url
        }
    }

@app.post("/api/contact")
def send_contact_message(data: schemas.ContactMessageRequest, request: Request):
    client_ip = get_client_ip(request)
    if not check_rate_limit(f"contact_{client_ip}", max_requests=3, window_seconds=300):
        raise HTTPException(
            status_code=429,
            detail="Muitas mensagens enviadas recentemente. Para evitar sobrecarga, aguarde 5 minutos antes de enviar uma nova mensagem."
        )

    if len(data.mensagem.strip()) < 10:
        raise HTTPException(status_code=400, detail="Por favor, digite uma mensagem com pelo menos 10 caracteres.")

    categoria = data.categoria or "Dúvidas Acadêmicas"
    email_service.send_contact_form_email(
        nome=data.nome.strip(),
        email=data.email.strip(),
        categoria=categoria.strip(),
        assunto=data.assunto.strip(),
        mensagem=data.mensagem.strip()
    )

    return {
        "success": True,
        "message": "Sua mensagem foi enviada com sucesso para a coordenação do Palieduca/UFPB! Responderemos em breve."
    }

@app.post("/api/auth/login")
def login(request: Request, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    client_ip = request.client.host if request.client else "unknown"
    if not check_rate_limit(f"login_{client_ip}", max_requests=10, window_seconds=60):
        raise HTTPException(
            status_code=429,
            detail="Muitas tentativas de login consecutivas. Por segurança, aguarde 1 minuto para tentar novamente."
        )

    user = db.query(models.User).filter(models.User.email == form_data.username).first()
    if not user or not user.senha_hash or not auth.verify_password(form_data.password, user.senha_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email ou senha incorretos",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Se ainda não verificou o email (e não for dono/admin)
    if not user.email_verified and user.cargo == "aluno":
        # Reenvia código
        code = email_service.generate_verification_code()
        user.verification_code = code
        db.commit()
        email_service.send_verification_email(user.email, user.nome, code)
        
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Por favor, confirme seu e-mail antes de entrar. Um novo código de 6 dígitos foi enviado para a sua caixa de entrada."
        )
    
    access_token = auth.create_access_token(data={"sub": user.email, "role": user.cargo})
    
    return {
        "access_token": access_token, 
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "email": user.email,
            "nome": user.nome,
            "cargo": user.cargo,
            "foto_url": user.foto_url
        }
    }

@app.post("/api/auth/google")
def google_auth(token_data: schemas.GoogleToken, db: Session = Depends(get_db)):
    idinfo = auth.verify_google_token(token_data.token)
    if not idinfo:
        raise HTTPException(status_code=401, detail="Token do Google inválido ou expirado.")
    
    email = idinfo.get("email")
    nome = idinfo.get("name", "Usuário do Google")
    
    # Verifica se o usuário já existe
    user = db.query(models.User).filter(models.User.email == email).first()
    
    # Se não existir, cria o usuário automaticamente com email já verificado pelo Google
    if not user:
        user = models.User(
            email=email,
            nome=nome,
            senha_hash=None, # Não tem senha local
            cargo="aluno",
            email_verified=True,
            auth_provider="google",
            foto_url=idinfo.get("picture")
        )
        db.add(user)
        db.commit()
        db.refresh(user)
    else:
        # Se já existia, garante que está verificado e atualiza foto caso não tenha foto customizada
        if not user.email_verified:
            user.email_verified = True
        if not user.foto_url and idinfo.get("picture"):
            user.foto_url = idinfo.get("picture")
        db.commit()
        db.refresh(user)
        
    access_token = auth.create_access_token(data={"sub": user.email, "role": user.cargo})
    
    return {
        "access_token": access_token, 
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "email": user.email,
            "nome": user.nome,
            "cargo": user.cargo,
            "foto_url": user.foto_url
        }
    }

@app.get("/api/auth/me", response_model=schemas.UserResponse)
def read_users_me(current_user: models.User = Depends(get_current_user)):
    return current_user

@app.post("/api/auth/change-password")
def change_password(
    data: schemas.ChangePasswordRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Verifica a senha atual se tiver senha cadastrada
    if current_user.senha_hash:
        if not auth.verify_password(data.current_password, current_user.senha_hash):
            raise HTTPException(status_code=400, detail="Senha atual incorreta.")
    
    # Validação de intervalo de 1 semana (7 dias)
    if current_user.last_password_change:
        try:
            last_change = datetime.fromisoformat(current_user.last_password_change)
            diff = datetime.now() - last_change
            if diff.total_seconds() < 7 * 24 * 3600:
                remaining_days = max(1, 7 - diff.days)
                raise HTTPException(
                    status_code=400,
                    detail=f"Por segurança, você só pode alterar a sua senha uma vez a cada 1 semana. Próxima alteração disponível em aproximadamente {remaining_days} dia(s)."
                )
        except Exception:
            pass

    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="A nova senha deve conter pelo menos 6 caracteres.")

    # Atualiza a senha e salva a data da troca
    current_user.senha_hash = auth.get_password_hash(data.new_password)
    current_user.last_password_change = datetime.now().isoformat()
    db.commit()
    db.refresh(current_user)

    return {
        "message": "Senha alterada com sucesso!",
        "last_password_change": current_user.last_password_change
    }

class PhotoUrlRequest(BaseModel):
    foto_url: str

@app.post("/api/auth/profile-photo")
async def update_profile_photo(
    file: UploadFile = File(None),
    foto_url: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if file and file.filename:
        filename_ext = os.path.splitext(file.filename)[1].lower()
        if filename_ext not in ALLOWED_AVATAR_EXTENSIONS:
            raise HTTPException(
                status_code=400, 
                detail=f"Extensão de imagem não permitida ({filename_ext}). Use: {', '.join(sorted(ALLOWED_AVATAR_EXTENSIONS))}"
            )
            
        clean_name = sanitize_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        random_id = uuid.uuid4().hex[:6]
        unique_filename = f"avatar_{current_user.id}_{timestamp}_{random_id}{filename_ext}"
        
        os.makedirs("static/uploads", exist_ok=True)
        file_path = os.path.join("static/uploads", unique_filename)
        
        file_size = 0
        with open(file_path, "wb") as buffer:
            while True:
                chunk = await file.read(512 * 1024) # 512 KB chunks
                if not chunk:
                    break
                file_size += len(chunk)
                if file_size > MAX_AVATAR_FILE_SIZE:
                    buffer.close()
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    raise HTTPException(status_code=413, detail="A foto de perfil excede o limite máximo permitido de 5MB.")
                buffer.write(chunk)
            
        current_user.foto_url = f"/static/uploads/{unique_filename}"
    elif foto_url:
        current_user.foto_url = foto_url.strip()
    else:
        raise HTTPException(status_code=400, detail="Nenhum arquivo ou URL de imagem fornecida.")
        
    db.commit()
    db.refresh(current_user)
    return {
        "message": "Foto de perfil atualizada com sucesso!",
        "foto_url": current_user.foto_url,
        "user": {
            "id": current_user.id,
            "email": current_user.email,
            "nome": current_user.nome,
            "cargo": current_user.cargo,
            "foto_url": current_user.foto_url
        }
    }

@app.delete("/api/auth/profile-photo")
def delete_profile_photo(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    current_user.foto_url = None
    db.commit()
    db.refresh(current_user)
    return {
        "message": "Foto de perfil removida.",
        "user": {
            "id": current_user.id,
            "email": current_user.email,
            "nome": current_user.nome,
            "cargo": current_user.cargo,
            "foto_url": None
        }
    }

@app.get("/api/modules", response_model=list[schemas.ModuleResponse])
def get_modules(db: Session = Depends(get_db)):
    return db.query(models.Module).all()

@app.put("/api/modules/{module_id}", response_model=schemas.ModuleResponse)
def update_module(
    module_id: int, 
    module_update: schemas.ModuleUpdate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão para editar conteúdo")
        
    db_module = db.query(models.Module).filter(models.Module.id == module_id).first()
    if not db_module:
        raise HTTPException(status_code=404, detail="Módulo não encontrado")
        
    db_module.title = module_update.title
    db_module.description = module_update.description
    db_module.image_url = module_update.image_url
    
    db.commit()
    db.refresh(db_module)
    return db_module

@app.get("/api/pages/{page_name}", response_model=schemas.PageContentResponse)
def get_page_content(page_name: str, db: Session = Depends(get_db)):
    page = db.query(models.PageContent).filter(models.PageContent.page_name == page_name).first()
    if not page:
        return schemas.PageContentResponse(id=0, page_name=page_name, content="")
    return page

@app.get("/api/pages/{page_name}/edit", response_model=schemas.PageContentResponse)
def get_page_content_edit(page_name: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão para editar conteúdo")
    
    page = db.query(models.PageContent).filter(models.PageContent.page_name == page_name).first()
    if not page:
        return schemas.PageContentResponse(id=0, page_name=page_name, content="")
    return page

@app.put("/api/pages/{page_name}/draft", response_model=schemas.PageContentResponse)
def update_page_draft(
    page_name: str, 
    page_update: schemas.PageContentUpdate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão para editar conteúdo")
        
    page = db.query(models.PageContent).filter(models.PageContent.page_name == page_name).first()
    if not page:
        page = models.PageContent(page_name=page_name, content="", draft_content=page_update.draft_content)
        if page_update.meta_title is not None: page.meta_title = page_update.meta_title
        if page_update.meta_description is not None: page.meta_description = page_update.meta_description
        if page_update.slug is not None: page.slug = page_update.slug
        db.add(page)
    else:
        if page_update.draft_content is not None: page.draft_content = page_update.draft_content
        if page_update.meta_title is not None: page.meta_title = page_update.meta_title
        if page_update.meta_description is not None: page.meta_description = page_update.meta_description
        if page_update.slug is not None: page.slug = page_update.slug
        
    db.commit()
    db.refresh(page)
    return page

@app.post("/api/pages/{page_name}/publish", response_model=schemas.PageContentResponse)
def publish_page_content(
    page_name: str, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão para editar conteúdo")
        
    page = db.query(models.PageContent).filter(models.PageContent.page_name == page_name).first()
    if not page or not page.draft_content:
        raise HTTPException(status_code=400, detail="Não há rascunho para publicar")
        
    # Salva no histórico de versões
    revision = models.PageRevision(
        page_name=page_name,
        content=page.content,
        author_name=current_user.nome,
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        description="Publicação via Editor Visual"
    )
    db.add(revision)
    
    # Atualiza conteúdo ao vivo e limpa rascunho
    page.content = page.draft_content
    page.draft_content = None
    
    db.commit()
    db.refresh(page)
    return page

@app.get("/api/pages/{page_name}/revisions", response_model=list[schemas.PageRevisionResponse])
def get_page_revisions(page_name: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    return db.query(models.PageRevision).filter(models.PageRevision.page_name == page_name).order_by(models.PageRevision.id.desc()).all()

@app.post("/api/pages/{page_name}/revisions", response_model=schemas.PageRevisionResponse)
def create_page_revision(
    page_name: str,
    revision_data: schemas.PageRevisionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    author = revision_data.author_name or current_user.nome or "Editor"
    revision = models.PageRevision(
        page_name=page_name,
        content=revision_data.content,
        author_name=author,
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        description=revision_data.description or "Ponto de restauração manual"
    )
    db.add(revision)
    db.commit()
    db.refresh(revision)
    return revision

@app.delete("/api/pages/{page_name}/revisions/{revision_id}")
def delete_page_revision(
    page_name: str,
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    revision = db.query(models.PageRevision).filter(
        models.PageRevision.id == revision_id, 
        models.PageRevision.page_name == page_name
    ).first()
    if not revision:
        raise HTTPException(status_code=404, detail="Revisão não encontrada")
    
    db.delete(revision)
    db.commit()
    return {"message": "Revisão removida com sucesso"}

@app.post("/api/pages/{page_name}/revisions/{revision_id}/restore", response_model=schemas.PageContentResponse)
def restore_page_revision(
    page_name: str, 
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
        
    revision = db.query(models.PageRevision).filter(models.PageRevision.id == revision_id, models.PageRevision.page_name == page_name).first()
    if not revision:
        raise HTTPException(status_code=404, detail="Revisão não encontrada")
        
    page = db.query(models.PageContent).filter(models.PageContent.page_name == page_name).first()
    if page:
        page.draft_content = revision.content
        db.commit()
        db.refresh(page)
        return page
    raise HTTPException(status_code=404, detail="Página não encontrada")

# ================================
# CMS v1 Routes (used by ModuleContentEditor)
# ================================

@app.get("/api/v1/cms/pages")
def list_all_cms_pages(db: Session = Depends(get_db)):
    pages = db.query(models.PageContent).all()
    return [{"id": p.id, "page_name": p.page_name, "content": p.content, "meta_title": p.meta_title, "meta_description": p.meta_description, "slug": p.slug} for p in pages]

@app.get("/api/v1/cms/pages/{page_name}")
def get_cms_page(page_name: str, db: Session = Depends(get_db)):
    page = db.query(models.PageContent).filter(models.PageContent.page_name == page_name).first()
    if not page:
        return {"id": 0, "page_name": page_name, "content": "[]", "draft_content": None}
    return page

@app.put("/api/v1/cms/pages/{page_name}")
def update_cms_page(
    page_name: str,
    page_update: schemas.PageContentUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")

    page = db.query(models.PageContent).filter(models.PageContent.page_name == page_name).first()
    if not page:
        page = models.PageContent(
            page_name=page_name,
            content=page_update.content or "[]",
            draft_content=page_update.draft_content
        )
        db.add(page)
    else:
        if page_update.content is not None:
            page.content = page_update.content
        if page_update.draft_content is not None:
            page.draft_content = page_update.draft_content

    # Save a revision for history
    revision = models.PageRevision(
        page_name=page_name,
        content=page.content,
        author_name=current_user.nome,
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        description=page_update.slug or "Atualização via CMS"
    )
    db.add(revision)

    db.commit()
    db.refresh(page)
    return page

@app.post("/api/v1/cms/revisions/{revision_id}/restore")
def restore_cms_revision(
    revision_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")

    revision = db.query(models.PageRevision).filter(models.PageRevision.id == revision_id).first()
    if not revision:
        raise HTTPException(status_code=404, detail="Revisão não encontrada")

    page = db.query(models.PageContent).filter(models.PageContent.page_name == revision.page_name).first()
    if not page:
        raise HTTPException(status_code=404, detail="Página não encontrada")

    page.content = revision.content
    page.draft_content = revision.content
    db.commit()
    db.refresh(page)
    return page

# ================================
# PageEditor: PUT /api/pages/{page_name} for saving content directly
# ================================

@app.put("/api/pages/{page_name}", response_model=schemas.PageContentResponse)
def update_page_content_direct(
    page_name: str,
    page_update: schemas.PageContentUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão para editar conteúdo")

    page = db.query(models.PageContent).filter(models.PageContent.page_name == page_name).first()
    if not page:
        page = models.PageContent(
            page_name=page_name,
            content=page_update.content or "[]",
            draft_content=page_update.draft_content
        )
        db.add(page)
    else:
        if page_update.content is not None:
            page.content = page_update.content
        if page_update.draft_content is not None:
            page.draft_content = page_update.draft_content

    db.commit()
    db.refresh(page)
    return page

@app.post("/api/chat")
async def chat(request: ChatRequest):
    groq_api_key = os.getenv("GROQ_API_KEY")
    if not groq_api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY não configurada no servidor backend.")

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {groq_api_key}"
    }

    # Prepends the system prompt to the conversation history incoming from frontend
    ai_messages = [SYSTEM_PROMPT] + [msg.model_dump() for msg in request.messages]

    payload = {
        # Modelo oficial de última geração, raciocínio avançado e ultrarrápido na Groq (100% Gratuito)
        "model": "qwen/qwen3.6-27b",
        "messages": ai_messages,
        "temperature": 0.7,
        "max_tokens": 1500
    }

    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=35.0
            )
            response.raise_for_status()
            data = response.json()
            raw_reply = data.get("choices", [{}])[0].get("message", {}).get("content")
            if not raw_reply:
                raise ValueError("Resposta inválida da API do Groq.")
            
            # Limpa blocos de raciocínio <think> se existirem para entregar resposta final limpa ao aluno
            reply = re.sub(r'<think>.*?</think>', '', raw_reply, flags=re.DOTALL).strip()
            if not reply:
                reply = raw_reply.strip()

            return {"reply": reply}

        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

# ================================
# Agente IA Construtor de Páginas & Banco de Imagens
# ================================

CURATED_HEALTHCARE_IMAGES = [
    {
        "id": "med_1",
        "title": "Acolhimento e Escuta Terapêutica em Cuidados Paliativos",
        "category": "Cuidados Paliativos",
        "url": "https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=400&q=80",
        "author": "National Cancer Institute"
    },
    {
        "id": "med_2",
        "title": "Equipe Multiprofissional de Enfermagem em Hospital",
        "category": "Enfermagem",
        "url": "https://images.unsplash.com/photo-1584515933487-779824d29309?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1584515933487-779824d29309?auto=format&fit=crop&w=400&q=80",
        "author": "Online Marketing"
    },
    {
        "id": "med_3",
        "title": "Apoio e Conforto ao Paciente Idoso e Família",
        "category": "Idoso & Família",
        "url": "https://images.unsplash.com/photo-1581579438747-1dc8d17bbce4?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1581579438747-1dc8d17bbce4?auto=format&fit=crop&w=400&q=80",
        "author": "Dominik Lange"
    },
    {
        "id": "med_4",
        "title": "Estetoscópio e Avaliação Clínica de Sintomas",
        "category": "Equipamentos & Exames",
        "url": "https://images.unsplash.com/photo-1505751172876-fa1923c5c528?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1505751172876-fa1923c5c528?auto=format&fit=crop&w=400&q=80",
        "author": "Online Marketing"
    },
    {
        "id": "med_5",
        "title": "Presença Compassiva e Toque Terapêutico",
        "category": "Cuidados Paliativos",
        "url": "https://images.unsplash.com/photo-1516574187841-cb9cc2ca948b?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1516574187841-cb9cc2ca948b?auto=format&fit=crop&w=400&q=80",
        "author": "Marcelo Leal"
    },
    {
        "id": "med_6",
        "title": "Comunicação de Notícias Difíceis e Tomada de Decisão",
        "category": "Comunicação",
        "url": "https://images.unsplash.com/photo-1579684385127-1ef15d508118?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1579684385127-1ef15d508118?auto=format&fit=crop&w=400&q=80",
        "author": "National Cancer Institute"
    },
    {
        "id": "med_7",
        "title": "Profissional de Enfermagem em Consulta Acolhedora",
        "category": "Enfermagem",
        "url": "https://images.unsplash.com/photo-1559839734-2b71ea197ec2?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1559839734-2b71ea197ec2?auto=format&fit=crop&w=400&q=80",
        "author": "Bruno Rodrigues"
    },
    {
        "id": "med_8",
        "title": "Mãos Dadas: Humanização e Cuidado Integral",
        "category": "Cuidados Paliativos",
        "url": "https://images.unsplash.com/photo-1532938911079-1b06ac7ceec7?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1532938911079-1b06ac7ceec7?auto=format&fit=crop&w=400&q=80",
        "author": "Hush Naidoo Jade Photography"
    },
    {
        "id": "med_9",
        "title": "Discussão de Bioética e Diretivas Antecipadas",
        "category": "Bioética",
        "url": "https://images.unsplash.com/photo-1454165804606-c3d57bc86b40?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1454165804606-c3d57bc86b40?auto=format&fit=crop&w=400&q=80",
        "author": "Scott Graham"
    },
    {
        "id": "med_10",
        "title": "Manejo Farmacológico e Controle de Sintomas",
        "category": "Medicamentos",
        "url": "https://images.unsplash.com/photo-1587854692152-cbe660dbde88?auto=format&fit=crop&w=1200&q=80",
        "thumb_url": "https://images.unsplash.com/photo-1587854692152-cbe660dbde88?auto=format&fit=crop&w=400&q=80",
        "author": "Laurynas Mereckas"
    }
]

TRANSLATIONS = {
    "cuidados paliativos": "palliative care patient",
    "paliativo": "palliative care",
    "enfermagem": "nurse nursing healthcare",
    "enfermeira": "nurse caring patient",
    "enfermeiro": "nurse hospital care",
    "medicina": "doctor healthcare hospital",
    "médico": "doctor medical examination",
    "hospital": "hospital room clinic",
    "idoso": "elderly senior patient care",
    "idosos": "elderly senior healthcare",
    "familia": "family patient hospital comfort",
    "família": "family comforting patient",
    "dor": "pain management patient care",
    "estetoscopio": "stethoscope doctor clinic",
    "estetoscópio": "stethoscope medical exam",
    "remedio": "medicine pills pharmacy",
    "remédio": "medicine pills healthcare",
    "medicamentos": "medication pharmacy healthcare",
    "bioetica": "medical ethics discussion",
    "bioética": "medical ethics doctor",
    "comunicacao": "doctor patient conversation empathy",
    "comunicação": "doctor patient empathy talk",
    "luto": "comforting grief support",
    "espiritualidade": "peaceful meditation hope",
    "anatomia": "human anatomy medical"
}

@app.get("/api/ai/search-images")
async def search_healthcare_images(
    q: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    source: Optional[str] = Query("all")
):
    results = []
    
    # 1. Filtra acervo curado
    curated = CURATED_HEALTHCARE_IMAGES
    if category and category.lower() != "todas":
        curated = [img for img in curated if img["category"].lower() == category.lower()]
    if q and q.strip():
        term = q.lower().strip()
        curated = [img for img in curated if term in img["title"].lower() or term in img["category"].lower()]
    
    results.extend(curated)
    
    # 2. Busca ao vivo em bancos externos (Unsplash / Openverse / Creative Commons)
    if q and q.strip() and source != "curated":
        raw_q = q.lower().strip()
        translated_q = TRANSLATIONS.get(raw_q, raw_q)
        try:
            async with httpx.AsyncClient(timeout=8.0) as client:
                res = await client.get(
                    f"https://api.openverse.org/v1/images/?q={translated_q}&page_size=20",
                    headers={"User-Agent": "Palieduca-App/1.0"}
                )
                if res.status_code == 200:
                    data = res.json()
                    for item in data.get("results", []):
                        img_url = item.get("url")
                        thumb = item.get("thumbnail") or img_url
                        if img_url:
                            results.append({
                                "id": f"ext_{item.get('id')}",
                                "title": item.get("title") or f"Imagem sobre {q}",
                                "category": "Unsplash / Web",
                                "url": img_url,
                                "thumb_url": thumb,
                                "author": item.get("creator") or "Unsplash / CC"
                            })
        except Exception as err:
            print("Erro ao buscar banco externo:", err)
            
    return results

class ResolveUrlRequest(BaseModel):
    url: str

@app.post("/api/ai/resolve-image-url")
async def resolve_image_url(req: ResolveUrlRequest):
    raw_url = req.url.strip()
    if not raw_url:
        raise HTTPException(status_code=400, detail="URL inválida")
        
    # Se já for link direto de imagem
    if any(ext in raw_url.lower() for ext in [".jpg", ".jpeg", ".png", ".webp", ".gif", "images.unsplash.com"]):
        return {"image_url": raw_url, "title": "Imagem Externa"}
        
    # Se for página do Unsplash (unsplash.com/pt-br/fotos/...)
    try:
        async with httpx.AsyncClient(timeout=8.0, follow_redirects=True) as client:
            res = await client.get(raw_url, headers={"User-Agent": "Mozilla/5.0"})
            if res.status_code == 200:
                og_match = re.search(r'<meta property=["\']og:image["\'] content=["\']([^"\']+)["\']', res.text)
                title_match = re.search(r'<title>([^<]+)</title>', res.text)
                if og_match:
                    title = title_match.group(1).split("|")[0].strip() if title_match else "Imagem Unsplash"
                    return {"image_url": og_match.group(1), "title": title}
    except Exception as e:
        print("Erro ao resolver URL:", e)
        
    return {"image_url": raw_url, "title": "Imagem Externa"}

AI_AGENT_BUILDER_SYSTEM_PROMPT = """Você é o Agente Arquiteto de Páginas e Aulas do Palieduca (UFPB).
Sua missão é transformar o pedido do professor em blocos visuais estruturados para o construtor visual de páginas.
Responda APENAS com um bloco de código JSON puro iniciado por ```json e finalizado por ```, SEM NENHUM COMENTÁRIO com // ou /* */ no JSON.

Estrutura JSON esperada:
```json
{
  "summary": "Resumo em 1 frase da aula montada",
  "blocks": [
    {
      "id": "hero_1",
      "type": "HeroBlock",
      "data": {
        "title": "Título Principal da Aula",
        "subtitle": "Subtítulo explicativo e cativante",
        "badgeText": "Módulo • Cuidados Paliativos",
        "primaryButtonText": "Iniciar Estudo",
        "imageUrl": "https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=1200&q=80",
        "align": "left"
      }
    },
    {
      "id": "cards_1",
      "type": "FeatureCardsBlock",
      "data": {
        "title": "Tópicos Fundamentais",
        "subtitle": "Conceitos-chave para a prática assistencial",
        "cards": [
          {
            "id": "c1",
            "icon_name": "HeartPulse",
            "iconColor": "#059669",
            "iconBg": "#ecfdf5",
            "badge": "Tópico 1",
            "title": "Acolhimento e Escuta",
            "description": "Explicação concisa e prática para o aluno."
          },
          {
            "id": "c2",
            "icon_name": "MessageSquare",
            "iconColor": "#d97706",
            "iconBg": "#fef3c7",
            "badge": "Tópico 2",
            "title": "Comunicação Empática",
            "description": "Técnicas de escuta ativa e diálogo humanizado."
          }
        ]
      }
    },
    {
      "id": "text_1",
      "type": "TextBlock",
      "data": {
        "htmlContent": "<h3>Fundamentação Teórica</h3><p>Explicação detalhada alinhada às diretrizes da ANCP e OMS...</p>",
        "align": "left"
      }
    },
    {
      "id": "quiz_1",
      "type": "QuizBlock",
      "data": {
        "title": "Quiz de Fixação: Teste seus Conhecimentos",
        "description": "Responda à questão abaixo com feedback imediato.",
        "questions": [
          {
            "id": "q1",
            "question": "Enunciado da questão sobre o tema abordado?",
            "options": [
              "Opção A incorreta",
              "Opção B correta e fundamentada",
              "Opção C incorreta",
              "Opção D incorreta"
            ],
            "correctOptionIndex": 1,
            "explanation": "Explicação pedagógica detalhada da resposta correta."
          }
        ]
      }
    }
  ]
}
```

Ícones válidos para os cards: HeartPulse, Stethoscope, Scale, MessageSquare, ShieldCheck, Users, Brain, Activity, Clock, Award, BookOpen, Smile, FileText.
IMPORTANTE: Não coloque nenhum comentário // dentro do código JSON.
"""

@app.post("/api/ai/generate-blocks", response_model=schemas.AIGenerateBlocksResponse)
async def generate_page_blocks_agent(
    request: schemas.AIGenerateBlocksRequest,
    current_user: models.User = Depends(get_current_user)
):
    allowed_roles = ["dona", "desenvolvedor", "administrador", "professor", "coordenador", "monitor", "suporte"]
    if current_user.cargo not in allowed_roles:
        raise HTTPException(status_code=403, detail="Sem permissão para utilizar o Agente Construtor.")

    groq_api_key = os.getenv("GROQ_API_KEY")
    if not groq_api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY não configurada no backend/.env")

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {groq_api_key}"
    }

    user_prompt = request.prompt
    if request.target_type and request.target_type != "full_page":
        user_prompt += f" (Foco: {request.target_type})"
    if request.context_module:
        user_prompt += f" (Módulo: {request.context_module})"

    payload = {
        "model": "qwen/qwen3.6-27b",
        "messages": [
            {"role": "system", "content": AI_AGENT_BUILDER_SYSTEM_PROMPT + "\nIMPORTANTE: Retorne a resposta estritamente como um objeto JSON válido."},
            {"role": "user", "content": user_prompt}
        ],
        "response_format": {"type": "json_object"},
        "temperature": 0.3,
        "max_tokens": 4096
    }

    async with httpx.AsyncClient(timeout=45.0) as client:
        try:
            response = await client.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers=headers,
                json=payload
            )
            
            if response.status_code == 401:
                raise HTTPException(
                    status_code=400, 
                    detail="Chave da Groq API inválida ou expirada. Atualize a GROQ_API_KEY no arquivo backend/.env."
                )
            
            if response.status_code != 200:
                err_text = response.text[:200]
                raise HTTPException(status_code=response.status_code, detail=f"Erro na Groq API: {err_text}")

            data = response.json()
            raw_reply = data.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
            
            # Limpa eventuais tags de pensamento <think> se presentes
            clean = re.sub(r"<think>[\s\S]*?</think>", "", raw_reply, flags=re.DOTALL).strip()
            
            # Remove blocos markdown caso o modelo tenha envolvido
            if clean.startswith("```"):
                clean = re.sub(r"^```(?:json)?\s*", "", clean)
                clean = re.sub(r"\s*```$", "", clean)

            # Limpeza de comentários (preservando URLs http e https)
            clean = re.sub(r"(?<!http:)(?<!https:)//.*", "", clean)
            clean = re.sub(r"/\*[\s\S]*?\*/", "", clean)
            clean = re.sub(r",\s*([\}\]])", r"\1", clean)

            first_brace = clean.find('{')
            last_brace = clean.rfind('}')
            if first_brace != -1 and last_brace != -1 and last_brace > first_brace:
                json_str = clean[first_brace:last_brace + 1]
            else:
                json_str = clean

            parsed_data = json.loads(json_str, strict=False)

            summary = parsed_data.get("summary", "Blocos gerados com sucesso pela IA!")
            generated_blocks = parsed_data.get("blocks", [])

            # Garante IDs únicos para cada bloco gerado
            for idx, b in enumerate(generated_blocks):
                unique_suffix = uuid.uuid4().hex[:6]
                b["id"] = f"{b.get('type', 'block').lower()}_{int(datetime.now().timestamp())}_{idx}_{unique_suffix}"

            return {
                "summary": summary,
                "blocks": generated_blocks
            }

        except HTTPException:
            raise
        except Exception as e:
            print("Erro ao processar resposta do Agente IA:", e)
            raise HTTPException(status_code=500, detail=f"Erro ao processar blocos com IA: {str(e)}")

@app.get("/api/modules/{module_slug}/resources", response_model=list[schemas.InteractiveResourceResponse])
def get_interactive_resources(module_slug: str, db: Session = Depends(get_db)):
    return db.query(models.InteractiveResource).filter(models.InteractiveResource.module_slug == module_slug).all()

@app.post("/api/modules/{module_slug}/resources", response_model=schemas.InteractiveResourceResponse)
def create_interactive_resource(
    module_slug: str,
    resource: schemas.InteractiveResourceCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    new_resource = models.InteractiveResource(
        module_slug=module_slug,
        type=resource.type,
        title=resource.title,
        content_json=resource.content_json
    )
    db.add(new_resource)
    db.commit()
    db.refresh(new_resource)
    return new_resource

@app.put("/api/resources/{resource_id}", response_model=schemas.InteractiveResourceResponse)
def update_interactive_resource(
    resource_id: int,
    resource_update: schemas.InteractiveResourceUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
        
    db_resource = db.query(models.InteractiveResource).filter(models.InteractiveResource.id == resource_id).first()
    if not db_resource:
        raise HTTPException(status_code=404, detail="Recurso não encontrado")
        
    db_resource.title = resource_update.title
    db_resource.content_json = resource_update.content_json
    db.commit()
    db.refresh(db_resource)
    return db_resource

# ==========================================
# ESPAÇO DE DÚVIDAS & COMENTÁRIOS DAS AULAS (COM MODERAÇÃO AUTOMÁTICA)
# ==========================================

@app.get("/api/modules/{module_slug}/comments")
def get_module_comments(module_slug: str, db: Session = Depends(get_db)):
    """
    Retorna todos os comentários e respostas aninhadas de um módulo.
    """
    all_comments = (
        db.query(models.ModuleComment)
        .filter(models.ModuleComment.module_slug == module_slug)
        .order_by(models.ModuleComment.is_pinned.desc(), models.ModuleComment.id.asc())
        .all()
    )

    parent_map = {}
    top_level = []

    for comment in all_comments:
        c_dict = {
            "id": comment.id,
            "module_slug": comment.module_slug,
            "user_id": comment.user_id,
            "author_name": comment.author_name,
            "author_role": comment.author_role,
            "author_avatar": comment.author_avatar,
            "content": comment.content,
            "created_at": comment.created_at,
            "is_pinned": comment.is_pinned,
            "likes_count": comment.likes_count,
            "parent_id": comment.parent_id,
            "replies": []
        }
        parent_map[comment.id] = c_dict

    for comment in all_comments:
        c_dict = parent_map[comment.id]
        if comment.parent_id and comment.parent_id in parent_map:
            parent_map[comment.parent_id]["replies"].append(c_dict)
        elif not comment.parent_id:
            top_level.append(c_dict)

    return top_level

@app.post("/api/modules/{module_slug}/comments")
def create_module_comment(
    module_slug: str,
    payload: schemas.CreateCommentRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Cria uma nova dúvida ou relato no módulo, passando pela verificação do Bot de Moderação Ética.
    """
    # 1. Verificação do Bot de Moderação Ética / Anti-Palavrões
    is_clean, mod_feedback = moderation_bot.check_content(payload.content)
    if not is_clean:
        raise HTTPException(
            status_code=400,
            detail=mod_feedback
        )

    # 2. Se for resposta a um comentário existente, valida se o parent existe
    if payload.parent_id:
        parent = db.query(models.ModuleComment).filter(models.ModuleComment.id == payload.parent_id).first()
        if not parent:
            raise HTTPException(status_code=404, detail="Comentário pai não encontrado.")

    # 3. Criação do comentário no banco
    now_str = datetime.now().strftime("%d/%m/%Y às %H:%M")
    new_comment = models.ModuleComment(
        module_slug=module_slug,
        user_id=current_user.id,
        author_name=current_user.nome,
        author_role=current_user.cargo,
        author_avatar=current_user.foto_url,
        content=payload.content.strip(),
        created_at=now_str,
        is_pinned=False,
        likes_count=0,
        parent_id=payload.parent_id
    )

    db.add(new_comment)
    db.commit()
    db.refresh(new_comment)

    return {
        "success": True,
        "message": "Comentário publicado com sucesso!",
        "comment": {
            "id": new_comment.id,
            "module_slug": new_comment.module_slug,
            "user_id": new_comment.user_id,
            "author_name": new_comment.author_name,
            "author_role": new_comment.author_role,
            "author_avatar": new_comment.author_avatar,
            "content": new_comment.content,
            "created_at": new_comment.created_at,
            "is_pinned": new_comment.is_pinned,
            "likes_count": new_comment.likes_count,
            "parent_id": new_comment.parent_id,
            "replies": []
        }
    }

@app.post("/api/comments/{comment_id}/like")
def like_module_comment(comment_id: int, db: Session = Depends(get_db)):
    comment = db.query(models.ModuleComment).filter(models.ModuleComment.id == comment_id).first()
    if not comment:
        raise HTTPException(status_code=404, detail="Comentário não encontrado.")

    comment.likes_count = (comment.likes_count or 0) + 1
    db.commit()
    return {"success": True, "likes_count": comment.likes_count}

@app.delete("/api/comments/{comment_id}")
def delete_module_comment(
    comment_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    comment = db.query(models.ModuleComment).filter(models.ModuleComment.id == comment_id).first()
    if not comment:
        raise HTTPException(status_code=404, detail="Comentário não encontrado.")

    # Autor ou Administradores / Professores podem excluir
    if comment.user_id != current_user.id and current_user.cargo not in ["dona", "desenvolvedor", "professor", "moderador"]:
        raise HTTPException(status_code=403, detail="Você não tem permissão para excluir este comentário.")

    # Exclui respostas filhas se houver
    db.query(models.ModuleComment).filter(models.ModuleComment.parent_id == comment_id).delete()
    db.delete(comment)
    db.commit()
    return {"success": True, "message": "Comentário excluído com sucesso."}

@app.delete("/api/resources/{resource_id}")
def delete_interactive_resource(
    resource_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
        
    db_resource = db.query(models.InteractiveResource).filter(models.InteractiveResource.id == resource_id).first()
    if not db_resource:
        raise HTTPException(status_code=404, detail="Recurso não encontrado")
        
    db.delete(db_resource)
    db.commit()
    return {"message": "Recurso deletado com sucesso"}

# ================================
# Biblioteca de Mídia
# ================================

@app.get("/api/media", response_model=list[schemas.MediaFileResponse])
def list_media_files(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    return db.query(models.MediaFile).order_by(models.MediaFile.id.desc()).all()

@app.post("/api/media/upload", response_model=schemas.MediaFileResponse)
async def upload_media_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
        
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Nenhum arquivo enviado.")

    filename_ext = os.path.splitext(file.filename)[1].lower()
    if filename_ext not in ALLOWED_MEDIA_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"Extensão de arquivo não permitida ({filename_ext}). Extensões aceitas: {', '.join(sorted(ALLOWED_MEDIA_EXTENSIONS))}"
        )

    # Sanitização do nome original e geração de caminho seguro contra Path Traversal
    clean_name = sanitize_filename(file.filename)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    random_id = uuid.uuid4().hex[:8]
    safe_filename = f"{timestamp}_{random_id}_{clean_name}"
    
    os.makedirs("static/uploads", exist_ok=True)
    file_path = os.path.join("static/uploads", safe_filename)
    
    # Leitura em chunks com limite de 25MB (Proteção contra DoS por estouro de memória/disco)
    file_size = 0
    with open(file_path, "wb") as buffer:
        while True:
            chunk = await file.read(1024 * 1024) # 1 MB chunks
            if not chunk:
                break
            file_size += len(chunk)
            if file_size > MAX_MEDIA_FILE_SIZE:
                buffer.close()
                if os.path.exists(file_path):
                    os.remove(file_path)
                raise HTTPException(status_code=413, detail="O arquivo excede o limite máximo permitido de 25MB.")
            buffer.write(chunk)
        
    # A URL que será devolvida pro frontend (caminho relativo)
    file_url = f"/static/uploads/{safe_filename}"
    
    # Salvar no banco
    media = models.MediaFile(
        filename=clean_name,
        file_url=file_url,
        uploaded_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )
    db.add(media)
    db.commit()
    db.refresh(media)
    return media

class DriveLinkRequest(BaseModel):
    url: str
    filename: str | None = None

@app.post("/api/media/drive-link", response_model=schemas.MediaFileResponse)
def add_google_drive_link(
    data: DriveLinkRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")

    url = data.url.strip()
    import re
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", url) or re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    if not match:
        raise HTTPException(status_code=400, detail="Link do Google Drive inválido. Use o link de compartilhamento da imagem.")

    file_id = match.group(1)
    cdn_url = f"https://lh3.googleusercontent.com/d/{file_id}"
    safe_filename = data.filename or f"GoogleDrive_{file_id[:8]}.jpg"

    media = models.MediaFile(
        filename=safe_filename,
        file_url=cdn_url,
        uploaded_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )
    db.add(media)
    db.commit()
    db.refresh(media)
    return media

@app.post("/api/media/workdrive-link", response_model=schemas.MediaFileResponse)
def add_zoho_workdrive_link(
    data: DriveLinkRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")

    url = data.url.strip()
    import re
    # Match Zoho WorkDrive file id (ex: /file/12345 or /embed/12345 or /download/12345)
    match = re.search(r"/(?:file|embed|download|open)/([a-zA-Z0-9_-]+)", url) or re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    
    if match:
        file_id = match.group(1)
        # Zoho WorkDrive embed / download link
        if "zohopublic" in url:
            direct_url = f"https://workdrive.zohopublic.com/download/{file_id}"
        elif "zohoexternal" in url:
            direct_url = f"https://workdrive.zohoexternal.com/download/{file_id}"
        else:
            direct_url = f"https://workdrive.zoho.com/download/{file_id}"
        safe_filename = data.filename or f"ZohoWorkDrive_{file_id[:8]}.jpg"
    else:
        # Se for link direto de imagem do Zoho
        direct_url = url
        safe_filename = data.filename or f"ZohoWorkDrive_{datetime.now().strftime('%H%M%S')}.jpg"

    media = models.MediaFile(
        filename=safe_filename,
        file_url=direct_url,
        uploaded_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )
    db.add(media)
    db.commit()
    db.refresh(media)
    return media

@app.delete("/api/media/{media_id}")
def delete_media_file(
    media_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    media = db.query(models.MediaFile).filter(models.MediaFile.id == media_id).first()
    if not media:
        raise HTTPException(status_code=404, detail="Arquivo de mídia não encontrado")
    
    if media.file_url and media.file_url.startswith("/static/uploads/"):
        disk_path = media.file_url.lstrip("/")
        if os.path.exists(disk_path):
            try:
                os.remove(disk_path)
            except Exception:
                pass

    db.delete(media)
    db.commit()
    return {"message": "Arquivo removido com sucesso"}

# ================================
# Progresso Granular & Atividades
# ================================

import json

def get_module_activity_ids(module_slug: str, db: Session) -> list[str]:
    """Retorna todos os IDs de atividades/blocos existentes dentro de um módulo."""
    activity_ids = []
    
    # 1. Busca blocos no PageContent (modulo_{slug})
    page = db.query(models.PageContent).filter(models.PageContent.page_name == f"modulo_{module_slug}").first()
    if page and page.content:
        try:
            blocks = json.loads(page.content)
            if isinstance(blocks, list):
                for b in blocks:
                    if isinstance(b, dict) and b.get("id"):
                        activity_ids.append(str(b["id"]))
        except Exception:
            pass

    # 2. Busca recursos interativos cadastrados para o módulo
    resources = db.query(models.InteractiveResource).filter(models.InteractiveResource.module_slug == module_slug).all()
    for r in resources:
        activity_ids.append(f"res_{r.id}")

    # 3. Se ainda não houver nenhum bloco, garante pelo menos o bloco padrão de introdução
    if not activity_ids:
        activity_ids.append(f"{module_slug}_intro")

    return list(set(activity_ids))

def check_and_send_completion_email(user: models.User, db: Session, force_send: bool = False) -> dict:
    """
    Verifica se o aluno completou 100% do curso e envia o e-mail oficial com o link do certificado.
    """
    if not force_send and user.completion_email_sent:
        return {"sent": False, "reason": "already_sent"}

    modules = db.query(models.Module).all()
    all_activity_ids = []
    for m in modules:
        all_activity_ids.extend(get_module_activity_ids(m.slug_id, db))
    total_activities = len(all_activity_ids)

    completed_records = db.query(models.UserActivityProgress).filter(
        models.UserActivityProgress.user_id == user.id,
        models.UserActivityProgress.completed == True
    ).all()
    completed_count = len(completed_records)

    is_eligible = (completed_count >= total_activities and total_activities > 0) or (user.cargo in ["dona", "desenvolvedor", "professor"])

    if not is_eligible:
        return {"sent": False, "reason": "not_eligible", "completed": completed_count, "total": total_activities}

    year = datetime.now().year
    certificate_code = f"PALI-{user.id:04d}-{year}-UFPB"
    verification_url = f"https://palieduca.com.br/validar/{certificate_code}"

    success = email_service.send_completion_congratulations_email(
        to_email=user.email,
        user_name=user.nome,
        certificate_code=certificate_code,
        verification_url=verification_url
    )

    if success or force_send:
        user.completion_email_sent = True
        db.commit()

    return {"sent": True, "certificate_code": certificate_code, "verification_url": verification_url}

def get_client_ip(request: Request) -> str:
    forwarded_for = request.headers.get("x-forwarded-for")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    real_ip = request.headers.get("x-real-ip")
    if real_ip:
        return real_ip.strip()
    if request.client and request.client.host:
        return request.client.host
    return "127.0.0.1"

@app.get("/api/progress", response_model=schemas.ActivityProgressResponse)
def get_user_progress(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Busca todas as atividades que o aluno já completou
    user_progress_records = db.query(models.UserActivityProgress).filter(
        models.UserActivityProgress.user_id == current_user.id,
        models.UserActivityProgress.completed == True
    ).all()
    
    completed_ids = set(r.activity_id for r in user_progress_records)
    
    modules = db.query(models.Module).all()
    module_progress = {}
    
    total_activities_sum = 0
    total_completed_sum = 0

    for mod in modules:
        act_ids = get_module_activity_ids(mod.slug_id, db)
        completed_in_mod = [aid for aid in act_ids if aid in completed_ids]
        
        mod_total = len(act_ids)
        mod_done = len(completed_in_mod)
        mod_pct = round((mod_done / mod_total) * 100) if mod_total > 0 else 0
        
        module_progress[mod.slug_id] = {
            "completed": mod_done,
            "total": mod_total,
            "percentage": mod_pct
        }
        
        total_activities_sum += mod_total
        total_completed_sum += mod_done

    overall_pct = round((total_completed_sum / total_activities_sum) * 100) if total_activities_sum > 0 else 0

    return {
        "completed_activities": list(completed_ids),
        "module_progress": module_progress,
        "overall_percentage": overall_pct,
        "total_completed": total_completed_sum,
        "total_activities": total_activities_sum
    }

@app.post("/api/progress/toggle")
def toggle_activity_progress(
    data: schemas.ActivityToggleRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    progress_entry = db.query(models.UserActivityProgress).filter(
        models.UserActivityProgress.user_id == current_user.id,
        models.UserActivityProgress.activity_id == data.activity_id
    ).first()

    if data.completed:
        if not progress_entry:
            progress_entry = models.UserActivityProgress(
                user_id=current_user.id,
                module_slug=data.module_slug,
                activity_id=data.activity_id,
                completed=True,
                completed_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            db.add(progress_entry)
        else:
            progress_entry.completed = True
            progress_entry.completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    else:
        if progress_entry:
            progress_entry.completed = False

    db.commit()

    # Se a atividade foi concluída, verifica se o aluno alcançou 100% para enviar e-mail de congratulações
    email_result = None
    if data.completed:
        email_result = check_and_send_completion_email(current_user, db, force_send=False)

    return {
        "message": "Progresso atualizado com sucesso!", 
        "activity_id": data.activity_id, 
        "completed": data.completed,
        "completion_email": email_result
    }

@app.post("/api/progress/send-certificate-email")
def send_certificate_email_endpoint(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Permite que o próprio aluno solicite o envio/reenvio do e-mail comemorativo do certificado a partir do seu perfil.
    """
    result = check_and_send_completion_email(current_user, db, force_send=True)
    if not result.get("sent"):
        raise HTTPException(
            status_code=400,
            detail="Você ainda não concluiu 100% de todas as atividades necessárias para a emissão do certificado."
        )
    return {
        "success": True,
        "message": f"E-mail comemorativo com o Certificado Oficial ({result.get('certificate_code')}) enviado com sucesso para {current_user.email}!",
        "certificate_code": result.get("certificate_code")
    }

# ================================
# Endpoints do Modo Visitante (IP & Navegador)
# ================================

@app.get("/api/guest/progress", response_model=schemas.ActivityProgressResponse)
def get_guest_progress(
    request: Request,
    guest_id: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    client_ip = get_client_ip(request)
    
    query = db.query(models.GuestActivityProgress).filter(models.GuestActivityProgress.completed == True)
    if guest_id:
        query = query.filter(
            (models.GuestActivityProgress.guest_id == guest_id) | 
            (models.GuestActivityProgress.ip_address == client_ip)
        )
    else:
        query = query.filter(models.GuestActivityProgress.ip_address == client_ip)
        
    guest_records = query.all()
    completed_ids = set(r.activity_id for r in guest_records)

    modules = db.query(models.Module).all()
    module_progress = {}
    total_activities_sum = 0
    total_completed_sum = 0

    for mod in modules:
        act_ids = get_module_activity_ids(mod.slug_id, db)
        completed_in_mod = [aid for aid in act_ids if aid in completed_ids]
        
        mod_total = len(act_ids)
        mod_done = len(completed_in_mod)
        mod_pct = round((mod_done / mod_total) * 100) if mod_total > 0 else 0
        
        module_progress[mod.slug_id] = {
            "completed": mod_done,
            "total": mod_total,
            "percentage": mod_pct
        }
        
        total_activities_sum += mod_total
        total_completed_sum += mod_done

    overall_pct = round((total_completed_sum / total_activities_sum) * 100) if total_activities_sum > 0 else 0

    return {
        "completed_activities": list(completed_ids),
        "module_progress": module_progress,
        "overall_percentage": overall_pct,
        "total_completed": total_completed_sum,
        "total_activities": total_activities_sum
    }

@app.post("/api/guest/progress/toggle")
def toggle_guest_progress(
    data: schemas.GuestActivityToggleRequest,
    request: Request,
    db: Session = Depends(get_db)
):
    client_ip = get_client_ip(request)
    user_agent = request.headers.get("user-agent", "")
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    progress_entry = db.query(models.GuestActivityProgress).filter(
        models.GuestActivityProgress.guest_id == data.guest_id,
        models.GuestActivityProgress.activity_id == data.activity_id
    ).first()

    if not progress_entry and client_ip:
        progress_entry = db.query(models.GuestActivityProgress).filter(
            models.GuestActivityProgress.ip_address == client_ip,
            models.GuestActivityProgress.activity_id == data.activity_id
        ).first()

    if data.completed:
        if not progress_entry:
            progress_entry = models.GuestActivityProgress(
                guest_id=data.guest_id,
                ip_address=client_ip,
                user_agent=user_agent,
                module_slug=data.module_slug,
                activity_id=data.activity_id,
                completed=True,
                completed_at=now_str,
                updated_at=now_str
            )
            db.add(progress_entry)
        else:
            progress_entry.guest_id = data.guest_id
            progress_entry.ip_address = client_ip
            progress_entry.user_agent = user_agent
            progress_entry.completed = True
            progress_entry.completed_at = now_str
            progress_entry.updated_at = now_str
    else:
        if progress_entry:
            progress_entry.completed = False
            progress_entry.updated_at = now_str

    db.commit()
    return {"message": "Progresso de visitante salvo com sucesso!", "activity_id": data.activity_id, "completed": data.completed}

# ================================
# Endpoints de Quizzes (Visitantes e Alunos)
# ================================

@app.post("/api/guest/quiz/answer")
def save_guest_quiz_answer(
    data: schemas.QuizAnswerSubmitRequest,
    request: Request,
    db: Session = Depends(get_db)
):
    client_ip = get_client_ip(request)
    guest_id = data.guest_id or "guest"
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    entry = db.query(models.GuestQuizAnswer).filter(
        models.GuestQuizAnswer.guest_id == guest_id,
        models.GuestQuizAnswer.block_id == data.block_id,
        models.GuestQuizAnswer.question_index == data.question_index
    ).first()

    if not entry and client_ip:
        entry = db.query(models.GuestQuizAnswer).filter(
            models.GuestQuizAnswer.ip_address == client_ip,
            models.GuestQuizAnswer.block_id == data.block_id,
            models.GuestQuizAnswer.question_index == data.question_index
        ).first()

    if not entry:
        entry = models.GuestQuizAnswer(
            guest_id=guest_id,
            ip_address=client_ip,
            module_slug=data.module_slug,
            block_id=data.block_id,
            question_index=data.question_index,
            selected_option=data.selected_option,
            is_correct=data.is_correct,
            answered_at=now_str
        )
        db.add(entry)
    else:
        entry.guest_id = guest_id
        entry.ip_address = client_ip
        entry.module_slug = data.module_slug or entry.module_slug
        entry.selected_option = data.selected_option
        entry.is_correct = data.is_correct
        entry.answered_at = now_str

    db.commit()
    return {"message": "Resposta salva com sucesso!", "is_correct": data.is_correct}

@app.get("/api/guest/quiz/answers")
def get_guest_quiz_answers(
    request: Request,
    guest_id: Optional[str] = Query(None),
    block_id: Optional[str] = Query(None),
    module_slug: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    client_ip = get_client_ip(request)
    query = db.query(models.GuestQuizAnswer)

    if guest_id:
        query = query.filter((models.GuestQuizAnswer.guest_id == guest_id) | (models.GuestQuizAnswer.ip_address == client_ip))
    else:
        query = query.filter(models.GuestQuizAnswer.ip_address == client_ip)

    if block_id:
        query = query.filter(models.GuestQuizAnswer.block_id == block_id)
    if module_slug:
        query = query.filter(models.GuestQuizAnswer.module_slug == module_slug)

    answers = query.all()
    return [
        {
            "block_id": a.block_id,
            "question_index": a.question_index,
            "selected_option": a.selected_option,
            "is_correct": a.is_correct,
            "answered_at": a.answered_at
        }
        for a in answers
    ]

@app.post("/api/quiz/answer")
def save_user_quiz_answer(
    data: schemas.QuizAnswerSubmitRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entry = db.query(models.UserQuizAnswer).filter(
        models.UserQuizAnswer.user_id == current_user.id,
        models.UserQuizAnswer.block_id == data.block_id,
        models.UserQuizAnswer.question_index == data.question_index
    ).first()

    if not entry:
        entry = models.UserQuizAnswer(
            user_id=current_user.id,
            module_slug=data.module_slug,
            block_id=data.block_id,
            question_index=data.question_index,
            selected_option=data.selected_option,
            is_correct=data.is_correct,
            answered_at=now_str
        )
        db.add(entry)
    else:
        entry.module_slug = data.module_slug or entry.module_slug
        entry.selected_option = data.selected_option
        entry.is_correct = data.is_correct
        entry.answered_at = now_str

    db.commit()
    return {"message": "Resposta salva com sucesso!", "is_correct": data.is_correct}

@app.get("/api/quiz/answers")
def get_user_quiz_answers(
    block_id: Optional[str] = Query(None),
    module_slug: Optional[str] = Query(None),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(models.UserQuizAnswer).filter(models.UserQuizAnswer.user_id == current_user.id)
    if block_id:
        query = query.filter(models.UserQuizAnswer.block_id == block_id)
    if module_slug:
        query = query.filter(models.UserQuizAnswer.module_slug == module_slug)

    answers = query.all()
    return [
        {
            "block_id": a.block_id,
            "question_index": a.question_index,
            "selected_option": a.selected_option,
            "is_correct": a.is_correct,
            "answered_at": a.answered_at
        }
        for a in answers
    ]

@app.post("/api/progress/sync-guest")
def sync_guest_progress(
    data: schemas.GuestSyncRequest,
    request: Request,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    client_ip = get_client_ip(request)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    synced_act_count = 0
    synced_quiz_count = 0

    # 1. Sincroniza atividades passadas no body
    for act_id in (data.completed_activities or []):
        exists = db.query(models.UserActivityProgress).filter(
            models.UserActivityProgress.user_id == current_user.id,
            models.UserActivityProgress.activity_id == act_id
        ).first()
        if not exists:
            db.add(models.UserActivityProgress(
                user_id=current_user.id,
                activity_id=act_id,
                completed=True,
                completed_at=now_str
            ))
            synced_act_count += 1
        elif not exists.completed:
            exists.completed = True
            exists.completed_at = now_str
            synced_act_count += 1

    # 2. Sincroniza atividades do GuestActivityProgress do guest_id ou IP
    guest_query = db.query(models.GuestActivityProgress).filter(models.GuestActivityProgress.completed == True)
    if data.guest_id:
        guest_query = guest_query.filter(
            (models.GuestActivityProgress.guest_id == data.guest_id) | 
            (models.GuestActivityProgress.ip_address == client_ip)
        )
    else:
        guest_query = guest_query.filter(models.GuestActivityProgress.ip_address == client_ip)
    
    for g_rec in guest_query.all():
        exists = db.query(models.UserActivityProgress).filter(
            models.UserActivityProgress.user_id == current_user.id,
            models.UserActivityProgress.activity_id == g_rec.activity_id
        ).first()
        if not exists:
            db.add(models.UserActivityProgress(
                user_id=current_user.id,
                module_slug=g_rec.module_slug,
                activity_id=g_rec.activity_id,
                completed=True,
                completed_at=g_rec.completed_at or now_str
            ))
            synced_act_count += 1
        elif not exists.completed:
            exists.completed = True
            synced_act_count += 1

    # 3. Sincroniza quizzes passados no body
    for q in (data.quiz_answers or []):
        b_id = q.get("block_id")
        q_idx = q.get("question_index")
        if b_id is not None and q_idx is not None:
            entry = db.query(models.UserQuizAnswer).filter(
                models.UserQuizAnswer.user_id == current_user.id,
                models.UserQuizAnswer.block_id == b_id,
                models.UserQuizAnswer.question_index == q_idx
            ).first()
            if not entry:
                db.add(models.UserQuizAnswer(
                    user_id=current_user.id,
                    module_slug=q.get("module_slug"),
                    block_id=b_id,
                    question_index=q_idx,
                    selected_option=q.get("selected_option", 0),
                    is_correct=q.get("is_correct", False),
                    answered_at=q.get("answered_at", now_str)
                ))
                synced_quiz_count += 1

    # 4. Sincroniza quizzes do GuestQuizAnswer
    guest_quiz_query = db.query(models.GuestQuizAnswer)
    if data.guest_id:
        guest_quiz_query = guest_quiz_query.filter(
            (models.GuestQuizAnswer.guest_id == data.guest_id) |
            (models.GuestQuizAnswer.ip_address == client_ip)
        )
    else:
        guest_quiz_query = guest_quiz_query.filter(models.GuestQuizAnswer.ip_address == client_ip)

    for g_q in guest_quiz_query.all():
        entry = db.query(models.UserQuizAnswer).filter(
            models.UserQuizAnswer.user_id == current_user.id,
            models.UserQuizAnswer.block_id == g_q.block_id,
            models.UserQuizAnswer.question_index == g_q.question_index
        ).first()
        if not entry:
            db.add(models.UserQuizAnswer(
                user_id=current_user.id,
                module_slug=g_q.module_slug,
                block_id=g_q.block_id,
                question_index=g_q.question_index,
                selected_option=g_q.selected_option,
                is_correct=g_q.is_correct,
                answered_at=g_q.answered_at or now_str
            ))
            synced_quiz_count += 1

    db.commit()
    return {
        "message": "Progresso e respostas de visitante sincronizados com sucesso!",
        "synced_activities": synced_act_count,
        "synced_quiz_answers": synced_quiz_count
    }


# ================================
# Validação Pública de Certificados UFPB
# ================================

@app.get("/api/certificates/validate/{code}", response_model=schemas.CertificateValidationResponse)
def validate_certificate(code: str, db: Session = Depends(get_db)):
    clean_code = code.strip().upper()
    
    # Formato esperado: PALI-0001-2026-UFPB ou PALI-1-2026-UFPB
    match = re.match(r"^PALI-(\d+)-(\d{4})-UFPB$", clean_code)
    if not match:
        return {
            "valid": False,
            "code": clean_code,
            "status_label": "CÓDIGO INVÁLIDO",
            "message": "O formato do código fornecido não corresponde ao padrão oficial de certificados da UFPB (ex: PALI-0001-2026-UFPB)."
        }

    user_id = int(match.group(1))
    year = int(match.group(2))

    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        return {
            "valid": False,
            "code": clean_code,
            "status_label": "NÃO ENCONTRADO",
            "message": "Nenhum estudante com este registro foi localizado na base de dados do Palieduca/UFPB."
        }

    # Calcula o total de atividades cadastradas em todos os módulos
    modules = db.query(models.Module).all()
    all_activity_ids = []
    for m in modules:
        all_activity_ids.extend(get_module_activity_ids(m.slug_id, db))
    total_activities = len(all_activity_ids)

    # Busca as atividades concluídas pelo usuário
    completed_records = db.query(models.UserActivityProgress).filter(
        models.UserActivityProgress.user_id == user.id,
        models.UserActivityProgress.completed == True
    ).all()
    completed_count = len(completed_records)

    is_eligible = (completed_count >= total_activities and total_activities > 0) or (user.cargo in ["dona", "desenvolvedor", "professor"])

    if not is_eligible:
        return {
            "valid": False,
            "code": clean_code,
            "student_name": user.nome,
            "student_id": user.id,
            "status_label": "EM ANDAMENTO / NÃO CONCLUÍDO",
            "message": f"O estudante completou {completed_count} de {total_activities} atividades. O certificado oficial só é válido após 100% de conclusão."
        }

    # Data da última conclusão ou ano
    latest_date_str = None
    for r in completed_records:
        if r.completed_at:
            if not latest_date_str or r.completed_at > latest_date_str:
                latest_date_str = r.completed_at

    return {
        "valid": True,
        "code": clean_code,
        "student_name": user.nome,
        "student_id": user.id,
        "course_name": "Cuidados Paliativos em Enfermagem",
        "workload_hours": 40,
        "institution": "Universidade Federal da Paraíba (UFPB)",
        "department": "Departamento de Enfermagem",
        "coordinator": "Prof.ª Patrícia Maria de Oliveira Andrade",
        "issue_date": latest_date_str or f"Ano de {year}",
        "issue_year": year,
        "status_label": "AUTENTICADO & VÁLIDO",
        "message": "Certificado Oficial emitido e autenticado com êxito pela Universidade Federal da Paraíba (UFPB)."
    }

# ================================
# Painel da Dona: Métricas & Gestão
# ================================

VALID_ROLES = ["dona", "desenvolvedor", "professor", "moderador", "suporte", "aluno"]

@app.get("/api/admin/metrics", response_model=schemas.AdminDashboardMetrics)
def get_admin_dashboard_metrics(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor", "professor", "moderador"]:
        raise HTTPException(status_code=403, detail="Sem permissão de acesso ao painel.")

    # Módulos e total de atividades da plataforma
    modules = db.query(models.Module).all()
    all_activity_ids = []
    module_activity_map = {}
    
    for mod in modules:
        act_ids = get_module_activity_ids(mod.slug_id, db)
        all_activity_ids.extend(act_ids)
        module_activity_map[mod.slug_id] = {
            "title": mod.title,
            "activities": act_ids
        }
    
    total_activities_count = len(all_activity_ids)
    all_activity_set = set(all_activity_ids)

    # Busca todos os usuários cadastrados
    all_db_users = db.query(models.User).all()
    
    students = [u for u in all_db_users if u.cargo == "aluno"]
    team_members = [u for u in all_db_users if u.cargo != "aluno"]

    def process_user_metric(u: models.User):
        s_records = db.query(models.UserActivityProgress).filter(
            models.UserActivityProgress.user_id == u.id,
            models.UserActivityProgress.completed == True
        ).all()
        
        s_completed_ids = set(r.activity_id for r in s_records if r.activity_id in all_activity_set)
        s_done = len(s_completed_ids)
        s_pct = round((s_done / total_activities_count) * 100) if total_activities_count > 0 else 0
        points = s_done * 10

        return {
            "id": u.id,
            "nome": u.nome,
            "email": u.email,
            "email_verified": u.email_verified,
            "cargo": u.cargo,
            "foto_url": u.foto_url,
            "completed_activities_count": s_done,
            "total_activities_count": total_activities_count,
            "progress_percentage": s_pct,
            "points": points,
            "is_certificate_eligible": s_pct >= 100 and s_done > 0,
            "s_completed_ids": s_completed_ids
        }

    all_users_metrics = [process_user_metric(u) for u in all_db_users]
    student_metrics = [m for m in all_users_metrics if m["cargo"] == "aluno"]

    # Estatísticas de distribuição dos alunos
    completed_students_count = len([m for m in student_metrics if m["progress_percentage"] >= 100 and m["completed_activities_count"] > 0])
    in_progress_students_count = len([m for m in student_metrics if 0 < m["progress_percentage"] < 100])
    not_started_students_count = len([m for m in student_metrics if m["progress_percentage"] == 0])

    # Estatísticas por módulo da trilha
    module_completion_totals = {mod.slug_id: 0 for mod in modules}
    for m in student_metrics:
        for mod in modules:
            mod_acts = module_activity_map[mod.slug_id]["activities"]
            mod_done = len([aid for aid in mod_acts if aid in m["s_completed_ids"]])
            if len(mod_acts) > 0 and mod_done == len(mod_acts):
                module_completion_totals[mod.slug_id] += 1

    # Remove campo temporário auxiliar
    for m in all_users_metrics:
        m.pop("s_completed_ids", None)

    # Ordena por pontos
    student_metrics.sort(key=lambda x: (x["points"], x["progress_percentage"]), reverse=True)
    all_users_metrics.sort(key=lambda x: (x["points"], x["progress_percentage"]), reverse=True)

    total_progress_sum = sum(m["progress_percentage"] for m in student_metrics)
    avg_progress = round(total_progress_sum / len(student_metrics)) if len(student_metrics) > 0 else 0

    module_stats = []
    for mod in modules:
        mod_acts_count = len(module_activity_map[mod.slug_id]["activities"])
        comp_count = module_completion_totals[mod.slug_id]
        comp_rate = round((comp_count / len(student_metrics)) * 100) if len(student_metrics) > 0 else 0
        module_stats.append({
            "slug": mod.slug_id,
            "title": mod.title,
            "activities_count": mod_acts_count,
            "completed_students": comp_count,
            "completion_rate": comp_rate
        })

    return {
        "total_students": len(student_metrics),
        "total_team_members": len(team_members),
        "total_modules": len(modules),
        "total_activities": total_activities_count,
        "average_progress_percentage": avg_progress,
        "status_distribution": {
            "completed": completed_students_count,
            "in_progress": in_progress_students_count,
            "not_started": not_started_students_count
        },
        "module_stats": module_stats,
        "students": student_metrics,
        "all_users": all_users_metrics
    }

@app.get("/api/admin/users")
def get_all_admin_users(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor", "professor"]:
        raise HTTPException(status_code=403, detail="Sem permissão.")

    users = db.query(models.User).all()
    return [{
        "id": u.id,
        "nome": u.nome,
        "email": u.email,
        "cargo": u.cargo,
        "email_verified": u.email_verified,
        "foto_url": u.foto_url,
        "auth_provider": u.auth_provider
    } for u in users]

@app.put("/api/admin/users/{user_id}/cargo")
def update_user_role(
    user_id: int,
    data: schemas.UserRoleUpdateRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Apenas a Dona ou Desenvolvedor podem alterar cargos.")

    new_role = data.cargo.lower().strip()
    if new_role not in VALID_ROLES:
        raise HTTPException(
            status_code=400, 
            detail=f"Cargo inválido. Escolha entre: {', '.join(VALID_ROLES)}"
        )

    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")

    # Regras de Cibersegurança e Proteção da Dona:
    # 1. A conta institucional principal da Dona nunca pode ser rebaixada
    if target_user.email == "patriciaandrade@palieduca.com.br" and new_role != "dona":
        raise HTTPException(status_code=403, detail="O cargo da Dona Proprietária não pode ser alterado.")

    # 2. Desenvolvedores não podem promover ninguém ao cargo de Dona (Apenas a Dona pode passar esse título)
    if current_user.cargo == "desenvolvedor" and new_role == "dona" and current_user.email != "patriciaandrade@palieduca.com.br":
        raise HTTPException(status_code=403, detail="Apenas a Dona pode conceder o cargo de Dona.")

    # 3. Desenvolvedores não podem alterar o cargo de outro usuário que já seja 'dona'
    if current_user.cargo == "desenvolvedor" and target_user.cargo == "dona":
        raise HTTPException(status_code=403, detail="Desenvolvedores não podem alterar cargos de Donas.")

    target_user.cargo = new_role
    db.commit()
    db.refresh(target_user)

    return {
        "message": f"Cargo de {target_user.nome} atualizado para '{new_role}' com sucesso!",
        "user_id": target_user.id,
        "novo_cargo": target_user.cargo
    }

@app.delete("/api/admin/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Apenas a Dona ou Desenvolvedor podem remover usuários.")

    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")

    # Regras de Cibersegurança e Proteção da Dona:
    # 1. Não permite apagar a conta principal da Dona ou usuários com cargo de Dona
    if target_user.email == "patriciaandrade@palieduca.com.br" or target_user.cargo == "dona":
        raise HTTPException(status_code=403, detail="Não é permitido remover a conta da Dona Proprietária.")

    # 2. Não permite apagar a si próprio pelo painel administrativo
    if target_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Você não pode excluir sua própria conta pelo painel de controle.")

    user_name = target_user.nome
    user_email = target_user.email

    # Remove o progresso e atividades do usuário no banco
    db.query(models.UserActivityProgress).filter(models.UserActivityProgress.user_id == user_id).delete()

    # Remove o usuário do banco de dados
    db.delete(target_user)
    db.commit()

    return {
        "message": f"Usuário {user_name} ({user_email}) removido com sucesso!",
        "deleted_user_id": user_id
    }

import io
import openpyxl
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi.responses import Response

@app.get("/api/admin/export-students-excel")
def export_students_excel(
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    user = None
    if token:
        payload = auth.decode_access_token(token)
        if payload and payload.get("sub"):
            user = db.query(models.User).filter(models.User.email == payload.get("sub")).first()
    elif authorization and authorization.startswith("Bearer "):
        jwt_token = authorization.split(" ")[1]
        payload = auth.decode_access_token(jwt_token)
        if payload and payload.get("sub"):
            user = db.query(models.User).filter(models.User.email == payload.get("sub")).first()

    if not user or user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=401, detail="Sem permissão ou não autenticado.")

    # 1. Coleta e processamento dos dados
    modules = db.query(models.Module).all()
    all_activity_ids = []
    module_activity_map = {}
    for mod in modules:
        act_ids = get_module_activity_ids(mod.slug_id, db)
        all_activity_ids.extend(act_ids)
        module_activity_map[mod.slug_id] = {
            "title": mod.title,
            "activities": act_ids
        }
    
    total_activities = len(all_activity_ids)
    all_activity_set = set(all_activity_ids)
    students = db.query(models.User).filter(models.User.cargo == "aluno").all()

    completed_count = 0
    in_progress_count = 0
    not_started_count = 0
    module_completion_totals = {mod.slug_id: 0 for mod in modules}
    student_rows = []

    for s in students:
        s_records = db.query(models.UserActivityProgress).filter(
            models.UserActivityProgress.user_id == s.id,
            models.UserActivityProgress.completed == True
        ).all()
        s_completed_ids = set(r.activity_id for r in s_records if r.activity_id in all_activity_set)
        s_done = len(s_completed_ids)
        s_pct = round((s_done / total_activities) * 100) if total_activities > 0 else 0
        points = s_done * 10

        if s_pct >= 100 and s_done > 0:
            completed_count += 1
        elif s_pct > 0:
            in_progress_count += 1
        else:
            not_started_count += 1

        for mod in modules:
            mod_acts = module_activity_map[mod.slug_id]["activities"]
            mod_done = len([aid for aid in mod_acts if aid in s_completed_ids])
            if len(mod_acts) > 0 and mod_done == len(mod_acts):
                module_completion_totals[mod.slug_id] += 1

        student_rows.append({
            "nome": s.nome,
            "email": s.email,
            "verified": "Sim" if s.email_verified else "Não",
            "points": points,
            "completed": s_done,
            "total": total_activities,
            "progress": s_pct,
            "certificate": "Apto (100%)" if s_pct >= 100 and s_done > 0 else "Em Andamento" if s_pct > 0 else "Não Iniciado"
        })

    student_rows.sort(key=lambda x: (x["points"], x["progress"]), reverse=True)

    # 2. Construção do Excel (.xlsx) com Estilos e Gráficos Nativos
    wb = openpyxl.Workbook()
    
    PRIMARY_COLOR = "2C5E55" # Verde Palieduca
    HEADER_FILL = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="4A7C72", end_color="4A7C72", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F7FAF8", end_color="F7FAF8", fill_type="solid")
    CARD_FILL = PatternFill(start_color="EAF2EF", end_color="EAF2EF", fill_type="solid")

    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="E0ECE8")
    BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="2C5E55")
    REGULAR_FONT = Font(name="Calibri", size=10)
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='D0DCD7'),
        right=Side(style='thin', color='D0DCD7'),
        top=Side(style='thin', color='D0DCD7'),
        bottom=Side(style='thin', color='D0DCD7')
    )

    # =========================================================================
    # ABA 1: 📊 Dashboard & Gráficos
    # =========================================================================
    ws_dash = wb.active
    ws_dash.title = "📊 Dashboard & Graficos"
    ws_dash.views.sheetView[0].showGridLines = True

    # Banner Superior
    ws_dash.merge_cells('A1:L2')
    top_cell = ws_dash['A1']
    top_cell.value = "  UNIVERSIDADE FEDERAL DA PARAÍBA (UFPB) • PALIEDUCA — PAINEL DA TURMA"
    top_cell.font = TITLE_FONT
    top_cell.fill = HEADER_FILL
    top_cell.alignment = Alignment(vertical="center")

    ws_dash.merge_cells('A3:L3')
    sub_cell = ws_dash['A3']
    sub_cell.value = f"  Relatório Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')} | Professora: Patricia Maria de Oliveira Andrade"
    sub_cell.font = SUBTITLE_FONT
    sub_cell.fill = SUBHEADER_FILL
    sub_cell.alignment = Alignment(vertical="center")

    # KPIs
    avg_prog = round(sum(s['progress'] for s in student_rows) / len(student_rows)) if student_rows else 0
    kpis = [
        ("A5:B6", "Total Alunos", len(students)),
        ("C5:D6", "Módulos", len(modules)),
        ("E5:F6", "Atividades", total_activities),
        ("G5:H6", "Média Geral", f"{avg_prog}%"),
        ("I5:L6", "Aptos Certificado", f"{completed_count} de {len(students)}")
    ]

    for merge_range, label, val in kpis:
        ws_dash.merge_cells(merge_range)
        first_cell_ref = merge_range.split(":")[0]
        ws_dash[first_cell_ref] = f"{label}: {val}"
        ws_dash[first_cell_ref].font = Font(name="Calibri", size=11, bold=True, color="2C5E55")
        ws_dash[first_cell_ref].fill = CARD_FILL
        ws_dash[first_cell_ref].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash[first_cell_ref].border = THIN_BORDER

    # Tabela 1: Distribuição de Desempenho
    ws_dash['A9'] = "Status de Desempenho da Turma"
    ws_dash['A9'].font = BOLD_FONT
    ws_dash['A10'] = "Status"
    ws_dash['B10'] = "Qtd Alunos"
    ws_dash['A10'].fill = SUBHEADER_FILL
    ws_dash['B10'].fill = SUBHEADER_FILL
    ws_dash['A10'].font = HEADER_FONT
    ws_dash['B10'].font = HEADER_FONT

    dist_data = [
        ("100% Concluido (Apto)", completed_count),
        ("Em Andamento (1-99%)", in_progress_count),
        ("Nao Iniciado (0%)", not_started_count)
    ]
    for idx, (status_lbl, qty) in enumerate(dist_data, start=11):
        ws_dash[f'A{idx}'] = status_lbl
        ws_dash[f'B{idx}'] = qty
        ws_dash[f'A{idx}'].font = REGULAR_FONT
        ws_dash[f'B{idx}'].font = REGULAR_FONT
        ws_dash[f'A{idx}'].border = THIN_BORDER
        ws_dash[f'B{idx}'].border = THIN_BORDER
        ws_dash[f'B{idx}'].alignment = Alignment(horizontal="center")

    # Gráfico de Pizza Nativo (Pie Chart)
    pie = PieChart()
    pie.title = "Distribuicao de Desempenho dos Alunos"
    pie_labels = Reference(ws_dash, min_col=1, min_row=11, max_row=13)
    pie_data = Reference(ws_dash, min_col=2, min_row=10, max_row=13)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    pie.width = 14
    pie.height = 7
    ws_dash.add_chart(pie, "D9")

    # Tabela 2: Conclusão por Módulo
    ws_dash['A16'] = "Taxa de Conclusao por Modulo"
    ws_dash['A16'].font = BOLD_FONT
    ws_dash['A17'] = "Modulo"
    ws_dash['B17'] = "Conclusao (%)"
    ws_dash['A17'].fill = SUBHEADER_FILL
    ws_dash['B17'].fill = SUBHEADER_FILL
    ws_dash['A17'].font = HEADER_FONT
    ws_dash['B17'].font = HEADER_FONT

    row_start_mod = 18
    for idx, mod in enumerate(modules, start=row_start_mod):
        comp_count = module_completion_totals[mod.slug_id]
        comp_rate = round((comp_count / len(students)) * 100) if students else 0
        ws_dash[f'A{idx}'] = mod.title
        ws_dash[f'B{idx}'] = comp_rate
        ws_dash[f'A{idx}'].font = REGULAR_FONT
        ws_dash[f'B{idx}'].font = REGULAR_FONT
        ws_dash[f'A{idx}'].border = THIN_BORDER
        ws_dash[f'B{idx}'].border = THIN_BORDER
        ws_dash[f'B{idx}'].alignment = Alignment(horizontal="center")
    row_end_mod = row_start_mod + len(modules) - 1

    # Gráfico de Barras Nativo (Bar Chart)
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Taxa de Conclusao por Modulo (%)"
    bar.y_axis.title = "Conclusao (%)"
    bar.x_axis.title = "Modulos"
    bar_data = Reference(ws_dash, min_col=2, min_row=17, max_row=row_end_mod)
    bar_labels = Reference(ws_dash, min_col=1, min_row=row_start_mod, max_row=row_end_mod)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_labels)
    bar.legend = None
    bar.width = 14
    bar.height = 7
    ws_dash.add_chart(bar, "D16")

    # =========================================================================
    # ABA 2: 👥 Lista de Alunos & Desempenho
    # =========================================================================
    ws_students = wb.create_sheet(title="👥 Alunos & Notas")
    ws_students.views.sheetView[0].showGridLines = True

    headers = [
        "Posicao Ranking", "Nome Completo do Aluno", "E-mail", 
        "E-mail Verificado", "Pontos (pts)", "Atividades Feitas", 
        "Total Atividades", "Progresso (%)", "Status do Certificado"
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_students.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws_students.row_dimensions[1].height = 26

    for r_idx, s in enumerate(student_rows, start=2):
        row_data = [
            f"#{r_idx - 1}", s["nome"], s["email"], s["verified"],
            s["points"], s["completed"], s["total"], f"{s['progress']}%", s["certificate"]
        ]
        is_even = (r_idx % 2 == 0)
        for col_idx, val in enumerate(row_data, start=1):
            c = ws_students.cell(row=r_idx, column=col_idx, value=val)
            c.font = REGULAR_FONT
            c.border = THIN_BORDER
            if is_even:
                c.fill = ZEBRA_FILL
            if col_idx in [1, 4, 5, 6, 7, 8, 9]:
                c.alignment = Alignment(horizontal="center")

    # Auto-ajuste de largura de colunas
    for ws in [ws_dash, ws_students]:
        for col in ws.columns:
            vals = [str(cell.value or '') for cell in col if cell.value is not None]
            max_len = max((len(v) for v in vals), default=10)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 3. Salva em memória e envia
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Relatorio_Turma_Palieduca_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/admin/export-students-csv")
def export_students_csv(
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    user = None
    if token:
        payload = auth.decode_access_token(token)
        if payload and payload.get("sub"):
            user = db.query(models.User).filter(models.User.email == payload.get("sub")).first()
    elif authorization and authorization.startswith("Bearer "):
        jwt_token = authorization.split(" ")[1]
        payload = auth.decode_access_token(jwt_token)
        if payload and payload.get("sub"):
            user = db.query(models.User).filter(models.User.email == payload.get("sub")).first()

    if not user or user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=401, detail="Sem permissão ou não autenticado.")

    modules = db.query(models.Module).all()
    all_activity_ids = []
    for mod in modules:
        all_activity_ids.extend(get_module_activity_ids(mod.slug_id, db))
    
    total_activities = len(all_activity_ids)
    all_activity_set = set(all_activity_ids)
    students = db.query(models.User).filter(models.User.cargo == "aluno").all()

    # CSV com BOM UTF-8 e seções para Google Sheets e Excel
    csv_lines = [
        "\ufeff=== RELATORIO GERAL DA TURMA - PALIEDUCA (UFPB) ===",
        f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')};Total Alunos: {len(students)};Modulos Ativos: {len(modules)};Total Atividades: {total_activities}",
        "",
        "=== LISTA DE ALUNOS E DESEMPENHO ===",
        "Ranking;Nome;E-mail;E-mail Verificado;Pontos (pts);Atividades Concluidas;Total Atividades;Progresso (%);Certificado Liberado"
    ]
    
    student_list = []
    for s in students:
        s_records = db.query(models.UserActivityProgress).filter(
            models.UserActivityProgress.user_id == s.id,
            models.UserActivityProgress.completed == True
        ).all()
        s_done = len([r.activity_id for r in s_records if r.activity_id in all_activity_set])
        s_pct = round((s_done / total_activities) * 100) if total_activities > 0 else 0
        points = s_done * 10
        verified_text = "Sim" if s.email_verified else "Nao"
        cert_text = "Sim" if s_pct >= 100 and s_done > 0 else "Nao"
        
        student_list.append((points, s_pct, s.nome, s.email, verified_text, s_done, total_activities, cert_text))

    student_list.sort(key=lambda x: (x[0], x[1]), reverse=True)

    for idx, item in enumerate(student_list, start=1):
        pts, pct, name, mail, ver, done, tot, cert = item
        csv_lines.append(f"#{idx};{name};{mail};{ver};{pts};{done};{tot};{pct}%;{cert}")

    csv_content = "\n".join(csv_lines)
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=Alunos_Palieduca_{datetime.now().strftime('%Y%m%d')}.csv"}
    )

# ================================
# Backup de Seguranca em 1 Clique
# ================================

@app.get("/api/admin/backup/export")
def export_system_backup(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")

    pages = db.query(models.PageContent).all()
    modules = db.query(models.Module).all()
    resources = db.query(models.InteractiveResource).all()

    backup_data = {
        "platform": "Palieduca",
        "exported_at": datetime.now().isoformat(),
        "exported_by": current_user.nome,
        "pages": [{"page_name": p.page_name, "content": p.content, "draft_content": p.draft_content, "meta_title": p.meta_title, "meta_description": p.meta_description, "slug": p.slug} for p in pages],
        "modules": [{"slug_id": m.slug_id, "title": m.title, "description": m.description, "icon_name": m.icon_name, "resources": m.resources, "image_url": m.image_url, "delay": m.delay} for m in modules],
        "interactive_resources": [{"module_slug": r.module_slug, "type": r.type, "title": r.title, "content_json": r.content_json} for r in resources]
    }

    return backup_data

@app.post("/api/admin/backup/restore")
def restore_system_backup(
    backup_data: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.cargo not in ["dona", "desenvolvedor"]:
        raise HTTPException(status_code=403, detail="Sem permissão")

    try:
        # Restaura páginas
        for p in backup_data.get("pages", []):
            page = db.query(models.PageContent).filter(models.PageContent.page_name == p["page_name"]).first()
            if page:
                page.content = p.get("content", "")
                page.draft_content = p.get("draft_content")
                page.meta_title = p.get("meta_title")
                page.meta_description = p.get("meta_description")
                page.slug = p.get("slug")
            else:
                db.add(models.PageContent(**p))

        # Restaura módulos
        for m in backup_data.get("modules", []):
            mod = db.query(models.Module).filter(models.Module.slug_id == m["slug_id"]).first()
            if mod:
                mod.title = m.get("title", mod.title)
                mod.description = m.get("description", mod.description)
                mod.icon_name = m.get("icon_name", mod.icon_name)
                mod.image_url = m.get("image_url", mod.image_url)
            else:
                db.add(models.Module(**m))

        db.commit()
        return {"message": "Backup restaurado com sucesso!"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao restaurar backup: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)